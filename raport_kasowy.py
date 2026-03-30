"""
Raport Kasowy – generator XLSX i PDF zgodny z importem do Symfonia F&K
======================================================================
Logika biznesowa:
  • Faktury (JPK_FA / KSeF):
      - sprzedaż (firma = sprzedawca)  → KP
      - zakup    (firma = nabywca)      → KW
  • Wyciąg bankowy (PDF) – tylko wybrane operacje:
      wypłata/wpłata bankomat, przelew wewnętrzny/własny/między kontami
      kwota ujemna (debet) → KP (zasilenie kasy)
      kwota dodatnia (credit) → KW (rozchód kasy do banku)
  • Lista płac / diety → KW
"""

from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from pathlib import Path
from typing import Literal, Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ── Rejestracja czcionek Unicode (polskie znaki) ─────────────────────────────
_FONT_DIRS = [
    Path(__file__).parent / "fonts",               # repo: fonts/ obok raport_kasowy.py
    Path(__file__).parent.parent / "fonts",        # fallback poziom wyżej
    Path("/usr/share/fonts/truetype/dejavu"),       # Linux systemowy
    Path("/usr/share/fonts/dejavu"),
    Path("/System/Library/Fonts"),                  # macOS
    Path("C:/Windows/Fonts"),                       # Windows lokalny
]

def _find_font(filename: str) -> str:
    for d in _FONT_DIRS:
        p = Path(d) / filename
        if p.exists():
            return str(p)
    return ""

def _register_fonts():
    regular = _find_font("DejaVuSans.ttf")
    if not regular:
        return False
    pdfmetrics.registerFont(TTFont("DejaVu", regular))
    # Brak Bold? Uzyj tej samej czcionki – polskie znaki wazniejsze niz pogrubienie
    bold = _find_font("DejaVuSans-Bold.ttf") or regular
    pdfmetrics.registerFont(TTFont("DejaVu-Bold", bold))
    pdfmetrics.registerFontFamily("DejaVu", normal="DejaVu", bold="DejaVu-Bold")
    return True

_FONTS_OK = _register_fonts()
_F        = "DejaVu"      if _FONTS_OK else "Helvetica"
_F_BOLD   = "DejaVu-Bold" if _FONTS_OK else "Helvetica-Bold"

# Ustaw globalny domyślny font ReportLab na DejaVu
if _FONTS_OK:
    from reportlab.lib.styles import getSampleStyleSheet as _gss
    from reportlab.rl_config import canvas_basefontname
    import reportlab.rl_config as _rlc
    _rlc.canvas_basefontname = _F
    # Nadpisz domyślne style raz na poziomie modułu
    _base_styles = _gss()
    for _s in _base_styles.byName.values():
        if hasattr(_s, "fontName") and "Helvetica" in (_s.fontName or ""):
            _s.fontName = _F
        if hasattr(_s, "bulletFontName") and "Helvetica" in (getattr(_s, "bulletFontName", "") or ""):
            _s.bulletFontName = _F

DocType = Literal["KP", "KW"]


# ============================================================================
# Model danych
# ============================================================================

@dataclass(order=True)
class KasaRecord:
    data: date
    typ: DocType
    numer_dokumentu: str
    kontrahent: str
    nip: str
    kwota: Decimal
    lp: int = field(default=0, compare=False)

    def kwota_str(self) -> str:
        q = self.kwota.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return str(q).replace(".", ",")

    def data_str(self) -> str:
        return self.data.strftime("%Y-%m-%d")


# ============================================================================
# Parsery XML – JPK_FA i KSeF
# ============================================================================

_GOTOWKA_RE = re.compile(r"got[oó]wk[ai]|cash|got\.", re.IGNORECASE | re.UNICODE)

# KSeF FA(3): FormaPlatnosci 1=gotówka, 2=karta, 3=bon, 4=czek, 5=kredyt, 6=przelew
_KSEF_GOTOWKA_CODES = {"1", "gotówka", "gotowka", "cash"}

_FORMA_FIELDS = ["P_19A", "FormaPlatnosci", "SposobPlatnosci"]


def _get_ns(root: ET.Element) -> tuple[str, dict]:
    m = re.match(r"\{([^}]+)\}", root.tag)
    uri = m.group(1) if m else ""
    return uri, {"tns": uri} if uri else {}


def _all_text(elem: ET.Element) -> str:
    return " ".join(
        t for node in elem.iter()
        for t in [node.text or "", node.tail or ""]
        if t.strip()
    )


def _find_first(elem: ET.Element, ns: dict, *xpaths: str) -> str:
    for xpath in xpaths:
        v = elem.findtext(xpath, namespaces=ns) or ""
        if v.strip():
            return v.strip()
    return ""


def _nip_clean(v: str) -> str:
    return re.sub(r"[^0-9]", "", v)


def _kwota_dec(v: str) -> Decimal:
    try:
        return abs(Decimal(v.replace(",", ".")))
    except InvalidOperation:
        return Decimal("0")


# ---------------------------------------------------------------------------
# Wyznaczenie kierunku faktury: KP (sprzedaż) vs KW (zakup)
# ---------------------------------------------------------------------------

def _kierunek_faktury(
    faktura: ET.Element,
    ns: dict,
    firma_nip: str,
    fmt: str,
) -> DocType:
    """
    Porównuje NIP firmy z NIP sprzedawcy i nabywcy na fakturze.
    JPK_FA v4:  P_5B = NIP sprzedawcy,  P_4B = NIP nabywcy
    JPK_FA old: P_5B lub Podmiot1 NIP   vs  P_3A
    KSeF:       Podmiot1/NIP = sprzedawca, Podmiot2/NIP = nabywca
    Jeśli nie można ustalić → domyślnie KP (sprzedaż).
    """
    if not firma_nip:
        return "KP"

    fn = _nip_clean(firma_nip)

    if fmt == "ksef":
        nip_sprzedawcy = _nip_clean(_find_first(
            faktura, ns,
            "tns:Podmiot1/tns:DaneIdentyfikacyjne/tns:NIP",
            ".//tns:Podmiot1//tns:NIP",
        ))
        nip_nabywcy = _nip_clean(_find_first(
            faktura, ns,
            "tns:Podmiot2/tns:DaneIdentyfikacyjne/tns:NIP",
            ".//tns:Podmiot2//tns:NIP",
        ))
    else:
        # JPK_FA v4: P_5B = sprzedawca, P_4B = nabywca
        nip_sprzedawcy = _nip_clean(_find_first(faktura, ns, "tns:P_5B", "P_5B"))
        nip_nabywcy    = _nip_clean(_find_first(faktura, ns, "tns:P_4B", "P_4B", "tns:P_3A", "P_3A"))

    if fn and nip_sprzedawcy and fn == nip_sprzedawcy:
        return "KP"   # firma jest sprzedawcą → wpływ gotówki
    if fn and nip_nabywcy and fn == nip_nabywcy:
        return "KW"   # firma jest nabywcą → rozchód gotówki

    return "KP"  # domyślnie sprzedaż


# ---------------------------------------------------------------------------
# Wyodrębnienie pól faktury wspólnych dla JPK_FA i KSeF
# ---------------------------------------------------------------------------

def _extract_faktura(
    faktura: ET.Element,
    ns: dict,
    fmt: str,
    firma_nip: str,
) -> Optional[KasaRecord]:

    # Data wystawienia
    data_txt = _find_first(
        faktura, ns,
        "tns:P_1", "P_1", "tns:Fa/tns:P_1", ".//tns:P_1",
    )
    try:
        dok_date = datetime.strptime(data_txt[:10], "%Y-%m-%d").date()
    except (ValueError, IndexError):
        return None

    # Numer faktury: JPK_FA v4 → P_2A; KSeF → Fa/P_2; starsze → P_2
    numer = _find_first(
        faktura, ns,
        "tns:P_2A", "P_2A",
        "tns:Fa/tns:P_2", ".//tns:P_2",
        "tns:P_2", "P_2",
        "tns:NrFaktury",
    ) or "?"

    # Kierunek: KP / KW
    kierunek = _kierunek_faktury(faktura, ns, firma_nip, fmt)

    # Kontrahent – zależy od kierunku
    if kierunek == "KP":
        # sprzedaż → kontrahent to NABYWCA
        kontrahent = _find_first(
            faktura, ns,
            "tns:P_3C", "P_3C",                                      # JPK_FA v4
            "tns:Podmiot2/tns:DaneIdentyfikacyjne/tns:Nazwa",
            ".//tns:Podmiot2//tns:Nazwa",                             # KSeF
            "tns:P_3B", "P_3B", "tns:NabywcaNazwa",                  # starsze
        )
        nip = _nip_clean(_find_first(
            faktura, ns,
            "tns:P_4B", "P_4B",
            "tns:Podmiot2/tns:DaneIdentyfikacyjne/tns:NIP",
            ".//tns:Podmiot2//tns:NIP",
            "tns:P_3A", "P_3A", "tns:NabywcaNIP",
        ))
    else:
        # zakup → kontrahent to SPRZEDAWCA
        kontrahent = _find_first(
            faktura, ns,
            "tns:P_3A", "P_3A",                                      # JPK_FA v4 sprzedawca
            "tns:Podmiot1/tns:DaneIdentyfikacyjne/tns:Nazwa",
            ".//tns:Podmiot1//tns:Nazwa",                             # KSeF
            "tns:P_3B", "P_3B",
        )
        nip = _nip_clean(_find_first(
            faktura, ns,
            "tns:P_5B", "P_5B",
            "tns:Podmiot1/tns:DaneIdentyfikacyjne/tns:NIP",
            ".//tns:Podmiot1//tns:NIP",
        ))

    # Kwota P_15
    kwota = _kwota_dec(_find_first(
        faktura, ns,
        "tns:P_15", "P_15", "tns:Fa/tns:P_15", ".//tns:P_15",
        ".//tns:KwotaNaleznosci", ".//tns:WartoscFaktury",
    ) or "0")

    return KasaRecord(
        data=dok_date,
        typ=kierunek,
        numer_dokumentu=numer,
        kontrahent=kontrahent,
        nip=nip,
        kwota=kwota,
    )


# ---------------------------------------------------------------------------
# parse_jpk_fa
# ---------------------------------------------------------------------------

def _is_gotowka_jpk(faktura: ET.Element, ns: dict) -> bool:
    for f in _FORMA_FIELDS:
        v = faktura.findtext(f"tns:{f}", namespaces=ns) or faktura.findtext(f) or ""
        if v and (_GOTOWKA_RE.search(v) or v.strip() in _KSEF_GOTOWKA_CODES):
            return True
    for xpath in (".//tns:FormaPlatnosci", ".//FormaPlatnosci",
                  "tns:Fa/tns:Platnosc/tns:FormaPlatnosci"):
        v = faktura.findtext(xpath, namespaces=ns) or ""
        if v.strip() in _KSEF_GOTOWKA_CODES or _GOTOWKA_RE.search(v):
            return True
    return bool(_GOTOWKA_RE.search(_all_text(faktura)))


def parse_jpk_fa(
    xml_path: str | Path,
    only_cash: bool = True,
    firma_nip: str = "",
) -> tuple[list[KasaRecord], dict]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    uri, ns = _get_ns(root)

    # NIP firmy z Podmiot1 jeśli nie podany
    if not firma_nip:
        firma_nip = _nip_clean(
            root.findtext(".//tns:Podmiot1//tns:NIP", namespaces=ns)
            or root.findtext(".//Podmiot1//NIP")
            or ""
        )

    faktury = (
        root.findall(".//tns:Faktura", ns)
        or root.findall("tns:Faktura", ns)
        or root.findall("Faktura")
    )

    records, skipped = [], 0
    has_forma = False

    for faktura in faktury:
        # sprawdź czy jest pole formy płatności
        forma_val = any(
            faktura.findtext(f"tns:{f}", namespaces=ns) or faktura.findtext(f)
            for f in _FORMA_FIELDS
        )
        if forma_val:
            has_forma = True

        if only_cash and not _is_gotowka_jpk(faktura, ns):
            skipped += 1
            continue

        rec = _extract_faktura(faktura, ns, "jpk_fa", firma_nip)
        if rec:
            records.append(rec)

    return records, {
        "total": len(records) + skipped,
        "cash": len(records),
        "skipped": skipped,
        "namespace": uri,
        "only_cash": only_cash,
        "has_forma_platnosci": has_forma,
        "firma_nip": firma_nip,
    }


# ---------------------------------------------------------------------------
# parse_ksef
# ---------------------------------------------------------------------------

def parse_ksef(
    xml_path: str | Path,
    only_cash: bool = True,
    firma_nip: str = "",
) -> tuple[list[KasaRecord], dict]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    uri, ns = _get_ns(root)

    local_root = root.tag.split("}")[-1].lower() if "}" in root.tag else root.tag.lower()
    if local_root == "faktura":
        faktury = [root]
    else:
        faktury = (
            root.findall(".//tns:Faktura", ns)
            or root.findall("tns:Faktura", ns)
            or root.findall("Faktura")
        )

    records, skipped = [], 0

    for faktura in faktury:
        forma = _find_first(
            faktura, ns,
            "tns:Fa/tns:Platnosc/tns:FormaPlatnosci",
            "Fa/Platnosc/FormaPlatnosci",
            ".//tns:FormaPlatnosci",
            ".//FormaPlatnosci",
        )
        is_cash = forma in _KSEF_GOTOWKA_CODES or bool(_GOTOWKA_RE.search(forma))
        if only_cash and not is_cash:
            skipped += 1
            continue

        rec = _extract_faktura(faktura, ns, "ksef", firma_nip)
        if rec:
            records.append(rec)

    return records, {
        "total": len(records) + skipped,
        "cash": len(records),
        "skipped": skipped,
        "namespace": uri,
        "only_cash": only_cash,
        "has_forma_platnosci": True,
        "firma_nip": firma_nip,
    }


# ---------------------------------------------------------------------------
# parse_xml_faktura – auto-detekcja formatu
# ---------------------------------------------------------------------------

def parse_xml_faktura(
    xml_path: str | Path,
    only_cash: bool = True,
    firma_nip: str = "",
) -> tuple[list[KasaRecord], dict, str]:
    """
    Auto-detekcja JPK_FA vs KSeF.
    KSeF: namespace 13775/12648/wzor/2025/2023/2021 lub root=<Faktura>
    """
    tree = ET.parse(xml_path)
    root = tree.getroot()
    tag   = root.tag.lower()
    local = root.tag.split("}")[-1].lower() if "}" in root.tag else root.tag.lower()

    ksef_hints = ["13775", "12648", "wzor/2025", "wzor/2023", "wzor/2021/11", "wzor/2022/01"]
    fmt = "ksef" if (any(h in tag for h in ksef_hints) or local == "faktura") else "jpk_fa"

    if fmt == "ksef":
        recs, diag = parse_ksef(xml_path, only_cash, firma_nip)
    else:
        recs, diag = parse_jpk_fa(xml_path, only_cash, firma_nip)

    return recs, diag, fmt


# ============================================================================
# Parser wyciągu bankowego (PDF)
# ============================================================================

# ---------------------------------------------------------------------------
# Santander (i inne banki) format linia transakcji:
#   DATA1 DATA2 OPIS_TRANSAKCJI  KWOTA  SALDO
#   np. "2026-02-26 2026-02-26 UZNANIE Przelew pomiędzy swoimi rachunkami 1.000,00 3.225,19"
#
# Logika kierunku:
#   "UZNANIE" w opisie → kwota wpłynęła na konto (credit) → gotówka WYSZŁA z kasy → KW
#   brak "UZNANIE"     → kwota opuściła konto  (debet)  → gotówka WESZŁA do kasy → KP
# ---------------------------------------------------------------------------

# Regex dopasowujący linię transakcji bankowej
# Grupy: (1) data, (2) opis, (3) kwota, (4) saldo
_WB_TX_RE = re.compile(
    r"^(\d{4}-\d{2}-\d{2})\s+\d{4}-\d{2}-\d{2}\s+(.+?)\s+"
    r"([\d.]+,\d{2})\s+([\d.]+,\d{2})\s*$"
)

# Alternatywny format daty DD.MM.YYYY
_WB_TX_RE2 = re.compile(
    r"^(\d{2}\.\d{2}\.\d{4})\s+\d{2}\.\d{2}\.\d{4}\s+(.+?)\s+"
    r"([\d.]+,\d{2})\s+([\d.]+,\d{2})\s*$"
)

# Operacje kasowe – jedyne które trafiają do raportu
_WB_CASH_RE = re.compile(
    r"wypłat[a-z]*\s+(?:w\s+|z\s+)?bankomat"
    r"|wpłat[a-z]*\s+(?:w\s+|do\s+)?bankomat"
    r"|ATM\s+(?:withdrawal|deposit|cash)"
    r"|transakcja\s+wewnętrzn"
    r"|przelew\s+wewnętrzn"
    r"|przelew\s+własn"
    r"|przelew\s+między\s+(?:swoimi\s+)?(?:kontami|rachunkami)"
    r"|przelew\s+pomiędzy\s+(?:swoimi\s+)?(?:kontami|rachunkami)"
    r"|operacja\s+między\s+kontami"
    r"|uznanie\s+przelew\s+pomiędzy",
    re.IGNORECASE | re.UNICODE,
)

# Czytelne opisy
_WB_OPIS_MAP = [
    (re.compile(r"wypłat.{0,10}bankomat",         re.I), "Wypłata z bankomatu"),
    (re.compile(r"wpłat.{0,10}bankomat",          re.I), "Wpłata do bankomatu"),
    (re.compile(r"ATM.{0,5}withdrawal",           re.I), "Wypłata z bankomatu"),
    (re.compile(r"ATM.{0,5}deposit",              re.I), "Wpłata do bankomatu"),
    (re.compile(r"transakcja\s+wewnętrzn",        re.I), "Transakcja wewnętrzna"),
    (re.compile(r"przelew\s+wewnętrzn",           re.I), "Przelew wewnętrzny"),
    (re.compile(r"przelew\s+własn",               re.I), "Przelew własny"),
    (re.compile(r"przelew.{0,10}między.{0,10}kontami",   re.I), "Przelew między kontami"),
    (re.compile(r"przelew.{0,10}pomiędzy.{0,15}rachunkami", re.I | re.U), "Przelew między rachunkami"),
    (re.compile(r"operacja.{0,10}między",         re.I), "Operacja między kontami"),
]

_DATE_PARSE_FMTS = ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%d/%m/%Y")


def _parse_wb_date(txt: str) -> Optional[date]:
    for fmt in _DATE_PARSE_FMTS:
        try:
            return datetime.strptime(txt.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _parse_wb_amount(txt: str) -> Decimal:
    """Parsuje polską kwotę: 1.000,20 → Decimal('1000.20')"""
    # usuń separatory tysięcy (kropki), zamień przecinek dziesiętny na kropkę
    clean = txt.replace(".", "").replace(",", ".")
    try:
        return abs(Decimal(clean))
    except InvalidOperation:
        return Decimal("0")


def _opis_wb(opis: str) -> str:
    for pattern, label in _WB_OPIS_MAP:
        if pattern.search(opis):
            return label
    return opis[:60]


def parse_bank_pdf(
    pdf_path: str | Path,
    nr_prefix: str = "BNK",
) -> list[KasaRecord]:
    """
    Przetwarza wyciąg bankowy (PDF) – format Santander i podobne.

    Obsługuje operacje kasowe:
      - Wypłata/wpłata w bankomacie
      - Transakcja wewnętrzna
      - Przelew wewnętrzny / własny / między kontami / pomiędzy rachunkami

    Logika KP/KW:
      OBCIĄŻENIA (brak słowa UZNANIE) → KP  (gotówka zasila kasę)
      UZNANIA    (słowo UZNANIE w opisie) → KW  (gotówka opuszcza kasę)
    """
    records: list[KasaRecord] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines = text.splitlines()

            for line in lines:
                # Dopasuj format linii transakcji
                m = _WB_TX_RE.match(line.strip()) or _WB_TX_RE2.match(line.strip())
                if not m:
                    continue

                date_txt, opis_raw, kwota_txt, _ = m.group(1), m.group(2), m.group(3), m.group(4)

                # Czy to operacja kasowa?
                if not _WB_CASH_RE.search(opis_raw):
                    continue

                dok_date = _parse_wb_date(date_txt)
                if not dok_date:
                    continue

                kwota = _parse_wb_amount(kwota_txt)

                # Kierunek: UZNANIE = wpływ na konto = gotówka WYSZŁA z kasy = KW
                is_uznanie = bool(re.match(r"uznanie", opis_raw.strip(), re.I))
                typ: DocType = "KW" if is_uznanie else "KP"

                seq   = len(records) + 1
                numer = f"{nr_prefix}/{dok_date.strftime('%Y%m')}/{seq:03d}"
                opis  = _opis_wb(opis_raw)

                records.append(KasaRecord(
                    data=dok_date,
                    typ=typ,
                    numer_dokumentu=numer,
                    kontrahent=opis,
                    nip="",
                    kwota=kwota,
                ))

    return records


# ============================================================================
# Procesor listy płac i diet
# ============================================================================

def process_payroll(
    entries: list[dict],
    pay_date: date,
    nr_prefix: str = "LP",
    collective: bool = False,
) -> list[KasaRecord]:
    """
    Przetwarza listę wypłat gotówkowych.

    Każdy element entries może zawierać:
        - 'nazwisko'     : str
        - 'kwota'        : str | float | Decimal
        - 'typ_wyplaty'  : 'wynagrodzenie' | 'dieta' | ''  (domyślnie 'wynagrodzenie')

    collective=True  → jeden zbiorczy KW
    collective=False → osobny KW per pracownik/dieta
    """
    records: list[KasaRecord] = []

    def _kwota(e: dict) -> Decimal:
        try:
            return abs(Decimal(str(e.get("kwota", 0))).quantize(Decimal("0.01")))
        except InvalidOperation:
            return Decimal("0")

    def _opis(e: dict, idx: int) -> str:
        nazwisko = e.get("nazwisko", f"Pracownik {idx}").strip()
        typ_w    = e.get("typ_wyplaty", "wynagrodzenie").lower()
        if typ_w == "dieta":
            return f"Dieta – {nazwisko}" if nazwisko else f"Dieta #{idx}"
        return nazwisko or f"Pracownik {idx}"

    if collective:
        total = sum(_kwota(e) for e in entries)
        # Jeśli mieszane typy – zbiorczy opis ogólny
        typy = {e.get("typ_wyplaty", "wynagrodzenie").lower() for e in entries}
        if typy == {"dieta"}:
            opis_zb = "Wypłata diet – lista"
            prefix_zb = "DT"
        elif typy == {"wynagrodzenie"} or not typy - {"wynagrodzenie", ""}:
            opis_zb = "Wypłata wynagrodzeń – lista płac"
            prefix_zb = nr_prefix
        else:
            opis_zb = "Wypłata wynagrodzeń i diet – lista"
            prefix_zb = nr_prefix

        records.append(KasaRecord(
            data=pay_date,
            typ="KW",
            numer_dokumentu=f"{prefix_zb}/{pay_date.strftime('%Y%m')}/ZB",
            kontrahent=opis_zb,
            nip="",
            kwota=total,
        ))
    else:
        wynagrodzenia = [(e, i) for i, e in enumerate(entries, 1)
                         if e.get("typ_wyplaty", "wynagrodzenie").lower() != "dieta"]
        diety         = [(e, i) for i, e in enumerate(entries, 1)
                         if e.get("typ_wyplaty", "").lower() == "dieta"]

        # Wynagrodzenia – prefix LP
        for seq, (e, orig_idx) in enumerate(wynagrodzenia, 1):
            records.append(KasaRecord(
                data=pay_date,
                typ="KW",
                numer_dokumentu=f"{nr_prefix}/{pay_date.strftime('%Y%m')}/{seq:03d}",
                kontrahent=_opis(e, orig_idx),
                nip="",
                kwota=_kwota(e),
            ))

        # Diety – prefix DT
        for seq, (e, orig_idx) in enumerate(diety, 1):
            records.append(KasaRecord(
                data=pay_date,
                typ="KW",
                numer_dokumentu=f"DT/{pay_date.strftime('%Y%m')}/{seq:03d}",
                kontrahent=_opis(e, orig_idx),
                nip="",
                kwota=_kwota(e),
            ))

    return records


# ============================================================================
# Główna klasa – Raport Kasowy
# ============================================================================

class RaportKasowy:
    def __init__(self, okres: str = ""):
        self.okres      = okres
        self._records:  list[KasaRecord] = []
        self.saldo_pocz: Decimal = Decimal("0")   # saldo początkowe (= konc. poprzedniego)
        self.saldo_konc: Decimal = Decimal("0")   # wyliczane przy eksporcie

    def dodaj_rekord(self, record: KasaRecord) -> "RaportKasowy":
        self._records.append(record)
        return self

    def _prepare(self) -> list[KasaRecord]:
        sorted_records = sorted(self._records)
        counters: dict[str, int] = {"KP": 0, "KW": 0}
        for i, r in enumerate(sorted_records, 1):
            r.lp = i
            counters[r.typ] += 1
            r.numer_dokumentu = (
                f"{r.typ}/{counters[r.typ]:03d}"
                f"/{r.data.strftime('%m')}/{r.data.strftime('%Y')}"
            )
        total_kp = sum(r.kwota for r in sorted_records if r.typ == "KP")
        total_kw = sum(r.kwota for r in sorted_records if r.typ == "KW")
        self.saldo_konc = self.saldo_pocz + total_kp - total_kw
        return sorted_records

    # -- eksport XLSX -----------------------------------------------------------

    def eksportuj_xlsx(self, sciezka: str | Path) -> Path:
        records  = self._prepare()
        wb = Workbook()
        ws = wb.active
        ws.title = "RaportKasowy"

        col_widths = [6, 14, 6, 22, 35, 14, 14]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        font_data    = Font(name="Arial", size=10)
        align_center = Alignment(horizontal="center", vertical="center")
        align_right  = Alignment(horizontal="right",  vertical="center")
        align_left   = Alignment(horizontal="left",   vertical="center")
        fill_kp      = PatternFill("solid", start_color="E8F5E9")
        fill_kw      = PatternFill("solid", start_color="FFF3E0")
        thin         = Side(style="thin", color="CCCCCC")
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)
        aligns       = [align_center, align_center, align_center,
                        align_left, align_left, align_center, align_right]

        for r in records:
            ws.append([r.lp, r.data_str(), r.typ, r.numer_dokumentu,
                        r.kontrahent, r.nip, r.kwota_str()])
            fill    = fill_kp if r.typ == "KP" else fill_kw
            row_num = ws.max_row
            for col_idx, aln in enumerate(aligns, 1):
                cell            = ws.cell(row=row_num, column=col_idx)
                cell.font       = font_data
                cell.alignment  = aln
                cell.fill       = fill
                cell.border     = border
            kwota_cell = ws.cell(row=row_num, column=7)
            try:
                kwota_cell.value          = float(str(r.kwota_str()).replace(",", "."))
                kwota_cell.number_format  = '#,##0.00_-'
            except Exception:
                pass

        # Wiersz saldo początkowe (przed danymi) – pusty Lp, typ "SP"
        ws.insert_rows(1)
        saldo_pocz_row = ["", "", "SP", "Saldo początkowe", "", "", ""]
        ws.cell(row=1, column=1).value = ""
        ws.cell(row=1, column=2).value = ""
        ws.cell(row=1, column=3).value = "SP"
        ws.cell(row=1, column=4).value = "Saldo początkowe"
        ws.cell(row=1, column=5).value = ""
        ws.cell(row=1, column=6).value = ""
        ws.cell(row=1, column=7).value = float(self.saldo_pocz)
        fill_sp = PatternFill("solid", start_color="E8EAF6")
        font_sp = Font(name="Arial", size=10, bold=True)
        for col in range(1, 8):
            c = ws.cell(row=1, column=col)
            c.fill   = fill_sp
            c.font   = font_sp
            c.border = border
        ws.cell(row=1, column=3).alignment = align_center
        ws.cell(row=1, column=7).alignment = align_right
        ws.cell(row=1, column=7).number_format = '#,##0.00_-'

        # Wiersz saldo końcowe (po danych)
        last = ws.max_row + 1
        ws.cell(row=last, column=3).value = "SK"
        ws.cell(row=last, column=4).value = "Saldo końcowe"
        ws.cell(row=last, column=7).value = float(self.saldo_konc)
        fill_sk = PatternFill("solid", start_color="E8EAF6")
        font_sk = Font(name="Arial", size=10, bold=True)
        for col in range(1, 8):
            c = ws.cell(row=last, column=col)
            c.fill   = fill_sk
            c.font   = font_sk
            c.border = border
        ws.cell(row=last, column=3).alignment = align_center
        ws.cell(row=last, column=7).alignment = align_right
        ws.cell(row=last, column=7).number_format = '#,##0.00_-'

        out = Path(sciezka)
        wb.save(out)
        return out

    # -- eksport PDF ------------------------------------------------------------

    def eksportuj_pdf(self, sciezka: str | Path) -> Path:
        records = self._prepare()
        out     = Path(sciezka)

        doc = SimpleDocTemplate(
            str(out), pagesize=landscape(A4),
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=2*cm, bottomMargin=2*cm,
        )

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "rk_title",
            fontName=_F_BOLD, fontSize=14,
            textColor=colors.HexColor("#1a237e"), spaceAfter=4,
            leading=18,
        )
        sub_style = ParagraphStyle(
            "rk_sub",
            fontName=_F, fontSize=9,
            textColor=colors.gray, spaceAfter=10,
            leading=12,
        )

        story   = []
        story.append(Paragraph("RAPORT KASOWY", title_style))
        okres_txt = f"Okres: {self.okres}    " if self.okres else ""
        story.append(Paragraph(
            f"{okres_txt}Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            sub_style,
        ))

        headers         = ["Lp.", "Data", "Typ", "Nr dokumentu", "Kontrahent", "NIP", "Kwota (PLN)"]
        col_widths_pdf  = [1.2*cm, 2.5*cm, 1.2*cm, 4.5*cm, 9*cm, 3*cm, 3*cm]

        # Wiersz saldo początkowe
        data_table = [
            headers,
            ["", "", "SP", "Saldo początkowe", "", "", _fmt_dec(self.saldo_pocz)],
        ]
        total_kp = Decimal("0")
        total_kw = Decimal("0")

        for r in records:
            data_table.append([str(r.lp), r.data_str(), r.typ, r.numer_dokumentu,
                                r.kontrahent, r.nip, r.kwota_str()])
            if r.typ == "KP":
                total_kp += r.kwota
            else:
                total_kw += r.kwota

        data_table += [
            ["", "", "", "", "OBROTY", "KP:",    _fmt_dec(total_kp)],
            ["", "", "", "", "",       "KW:",    _fmt_dec(total_kw)],
            ["", "", "SK", "Saldo końcowe",   "", "", _fmt_dec(self.saldo_konc)],
        ]

        n  = len(data_table)
        ts = TableStyle([
            ("BACKGROUND",  (0, 0),     (-1, 0),     colors.HexColor("#1a237e")),
            ("TEXTCOLOR",   (0, 0),     (-1, 0),     colors.white),
            ("FONTNAME",    (0, 0),     (-1, 0),     _F_BOLD),
            ("FONTNAME",    (0, 1),     (-1, n-4),   _F),        # wiersze danych
            ("FONTSIZE",    (0, 0),     (-1, n-4),   8),
            ("VALIGN",      (0, 0),     (-1, -1),    "MIDDLE"),
            ("ALIGN",       (0, 0),     (-1, 0),     "CENTER"),
            ("ALIGN",       (0, 1),     (0, -1),     "CENTER"),
            ("ALIGN",       (1, 1),     (1, -1),     "CENTER"),
            ("ALIGN",       (2, 1),     (2, -1),     "CENTER"),
            ("ALIGN",       (5, 1),     (5, -1),     "CENTER"),
            ("ALIGN",       (6, 1),     (6, -1),     "RIGHT"),
            ("GRID",        (0, 0),     (-1, n-4),   0.3, colors.HexColor("#dddddd")),
            ("LINEABOVE",   (0, 0),     (-1, 0),     1.5, colors.HexColor("#1a237e")),
            ("LINEBELOW",   (0, 0),     (-1, 0),     1.5, colors.HexColor("#1a237e")),
            ("FONTNAME",    (0, n-3),   (-1, -1),    _F_BOLD),
            ("ALIGN",       (5, n-3),   (6, -1),     "RIGHT"),
            ("ALIGN",       (6, n-1),   (6, -1),     "RIGHT"),
            ("LINEABOVE",   (0, n-3),   (-1, n-3),   1, colors.HexColor("#1a237e")),
            ("BACKGROUND",  (0, n-3),   (-1, -1),    colors.HexColor("#e8eaf6")),
            # Saldo początkowe (wiersz 1) – wyróżnienie
            ("BACKGROUND",  (0, 1),     (-1, 1),     colors.HexColor("#e8eaf6")),
            ("FONTNAME",    (0, 1),     (-1, 1),     _F_BOLD),
            ("LINEBELOW",   (0, 1),     (-1, 1),     0.5, colors.HexColor("#1a237e")),
            *[("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f5f5f5"))
              for i in range(3, n-3, 2)],
            *[("TEXTCOLOR",  (2, i), (2, i),
               colors.HexColor("#1b5e20") if data_table[i][2] == "KP"
               else colors.HexColor("#bf360c") if data_table[i][2] == "KW"
               else colors.HexColor("#1a237e"))
              for i in range(1, n-3)],
        ])

        table = Table(data_table, colWidths=col_widths_pdf, repeatRows=1)
        table.setStyle(ts)
        story.append(table)

        story.append(Spacer(1, 1*cm))
        sign_table = Table(
            [["Sporządził:", "", "Zatwierdził:"],
             ["", "", ""],
             ["_________________________", "", "_________________________"]],
            colWidths=[7*cm, 5*cm, 7*cm],
        )
        sign_table.setStyle(TableStyle([
            ("FONTNAME",   (0, 0), (-1, -1), _F),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 1), (-1, 1),  14),
        ]))
        story.append(sign_table)
        doc.build(story)
        return out


def _fmt_dec(val: Decimal) -> str:
    return str(val.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)).replace(".", ",")
